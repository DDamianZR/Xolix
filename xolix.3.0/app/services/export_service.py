"""
Servicio de exportación: genera PDF y Excel a partir de datos del sistema.
PDF usa reportlab; Excel usa openpyxl.
"""
import io
from typing import List


def generar_pdf_casos(casos: list) -> bytes:
    """Genera un PDF con el listado de casos NNA."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Reporte de Casos NNA — Xolix", styles["Title"]))
    elements.append(Spacer(1, 12))

    headers = ["ID", "Nombre NNA", "Edad", "Estado", "Fecha Creación"]
    data = [headers]
    for c in casos:
        data.append([
            str(c.id),
            c.nna_nombre,
            str(c.nna_edad or "—"),
            c.estado.value if hasattr(c.estado, "value") else str(c.estado),
            c.fecha_creacion.strftime("%Y-%m-%d") if c.fecha_creacion else "—",
        ])

    t = Table(data, colWidths=[40, 200, 50, 90, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C3483")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EEF8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    return buf.getvalue()


def generar_excel_casos(casos: list) -> bytes:
    """Genera un Excel con el listado de casos NNA."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Casos NNA"

    header_fill = PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["ID", "Nombre NNA", "CURP", "Edad", "Género", "Nacionalidad", "Estado", "Fecha Creación"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for c in casos:
        ws.append([
            c.id,
            c.nna_nombre,
            c.nna_curp or "",
            c.nna_edad,
            c.nna_genero.value if c.nna_genero and hasattr(c.nna_genero, "value") else (c.nna_genero or ""),
            c.nna_nacionalidad or "",
            c.estado.value if hasattr(c.estado, "value") else str(c.estado),
            c.fecha_creacion.strftime("%Y-%m-%d") if c.fecha_creacion else "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_actores(actores: list) -> bytes:
    """Genera un Excel con el catálogo de actores."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Actores"

    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["ID", "Nombre", "Tipo", "Municipio", "Estado", "Teléfono", "Correo", "Activo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for a in actores:
        ws.append([
            a.id, a.nombre,
            a.tipo.value if hasattr(a.tipo, "value") else str(a.tipo),
            a.municipio or "", a.estado or "",
            a.telefono or "", a.correo or "",
            "Sí" if a.activo else "No",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf_diagnostico(diagnostico, caso_nna=None) -> bytes:
    """Genera un PDF con el resumen de un diagnóstico."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=50, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    title = f"Diagnóstico — {diagnostico.tipo.value.upper() if hasattr(diagnostico.tipo, 'value') else diagnostico.tipo}"
    elements.append(Paragraph(title, styles["Title"]))
    if caso_nna:
        elements.append(Paragraph(f"Caso NNA: {caso_nna.nna_nombre}", styles["Heading2"]))
    elements.append(Paragraph(f"Fecha: {diagnostico.fecha}", styles["Normal"]))
    elements.append(Paragraph(f"Observaciones: {diagnostico.observaciones or '—'}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    if diagnostico.derechos_vulnerados:
        elements.append(Paragraph("Derechos Vulnerados", styles["Heading2"]))
        dv_data = [["Derecho", "Severidad", "Recomendación"]]
        for dv in diagnostico.derechos_vulnerados:
            dv_data.append([
                str(dv.derecho_id),
                dv.severidad.value if hasattr(dv.severidad, "value") else str(dv.severidad),
                dv.recomendacion or "—",
            ])
        t = Table(dv_data, colWidths=[150, 80, 250])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C3483")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)

    doc.build(elements)
    return buf.getvalue()
